import os
from datetime import date, timedelta

class DataProviderClass:

    SEL_EXCEL_DATA_FILE = os.path.abspath(os.path.join("../testdata","testdata.xlsx"))
    SEL_JSON_DATA_FILE = os.path.abspath(os.path.join("../testdata","input.json"))

    API_EXCEL_DATA_FILE = os.path.abspath(os.path.join("../testdata","testdata-api.xlsx"))
    API_JSON_DATA_FILE = os.path.abspath(os.path.join("../testdata", "input-api.json"))

    TEST_DATA_DICT = {}

    @staticmethod
    def populate_data_from_local(type="sel"):
        if type == "sel":
            local_test_ajax_form_data = [
                {"username": "John Paul", "comments": "Hi this is John Paul"},
                {"username": "Swetha Shankar", "comments": "This works cool"},
                {"username": "Steve Waugh", "comments": "Hey Folks"},
                {"username": "Mark Waugh", "comments": "Hey Aussies"},
                {"username": "Sachin Ramesh", "comments": "Hello Indians"}
            ]

            local_test_jquery_date_data = [
                {"from_date": "2020-10-01", "to_date": "2020-10-01"},
                {"from_date": "2020-07-01", "to_date": "2020-11-01"},
                {"from_date": "2020-03-01", "to_date": "2020-02-01"}
            ]

            local_test_slider_range_data = [
                {"value": 50},
                {"value": 40}
            ]

            local_test_file_download_data = [
                {"content": "hello world"},
                {
                    "content": """Lorem Ipsum is simply dummy text of the printing and typesetting industry. Lorem Ipsum has been the industry's standard dummy text ever since the 1500s,when an unknown printer took a galley of type and scrambled it to make a type specimen book. It has survived not only five centuries, but also the leap into electronic typesetting, remaining essentially unchanged. It was popularised in the 1960s with the release of Letraset sheets containing Lorem Ipsum passages, and more recently with desktop p"""},
                {"content": "I’m designing a document and don’t want to get bogged down in what the text actually says."}
            ]

            DataProviderClass.TEST_DATA_DICT['ajax_form'] = local_test_ajax_form_data
            DataProviderClass.TEST_DATA_DICT['jquery_date'] = local_test_jquery_date_data
            DataProviderClass.TEST_DATA_DICT['slider_range'] = local_test_slider_range_data
            DataProviderClass.TEST_DATA_DICT['file_dw'] = local_test_file_download_data
        else:
            _checkin_date = date.today()
            _checkout_date = _checkin_date + timedelta(5)
            local_booking_data = {
                "firstname": "Tim",
                "lastname": "Cook",
                "totalprice": 1200,
                "depositpaid": True,
                "bookingdates": {
                    "checkin": str(_checkin_date),
                    "checkout": str(_checkout_date)
                },
                "additionalneeds": "Breakfast"
            }

            local_modified_data = {
                "firstname": "Timothy",
                "lastname": "Johnson",
                "totalprice": 1400,
                "depositpaid": True,
                "bookingdates": {
                    "checkin": str(_checkin_date),
                    "checkout": str(_checkout_date)
                },
                "additionalneeds": ["Breakfast", "Coffee"]
            }
            DataProviderClass.TEST_DATA_DICT['booking_data'] = local_booking_data
            DataProviderClass.TEST_DATA_DICT['modified_data'] = local_modified_data

    @staticmethod
    def populate_data_from_excel(type="sel"):
        import openpyxl
        try:
            if type == "sel":
                workbook = openpyxl.load_workbook(DataProviderClass.SEL_EXCEL_DATA_FILE)
            else:
                workbook = openpyxl.load_workbook(DataProviderClass.API_EXCEL_DATA_FILE)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                temp_list = []
                for i in range(2, sheet.max_row + 1):
                    temp_dict = {}
                    for j in range(1, sheet.max_column + 1):
                        temp_dict[sheet.cell(1, j).value] = sheet.cell(i, j).value
                    temp_list.append(temp_dict)
                DataProviderClass.TEST_DATA_DICT[sheet_name] = temp_list
        except FileNotFoundError:
            print("No Such File Found Exception")
        finally:
            workbook.close()

    @staticmethod
    def populate_data_from_json(type="sel"):
        import json
        try:
            if type == "sel":
                jfile = open(DataProviderClass.SEL_JSON_DATA_FILE,'r')
            else:
                jfile = open(DataProviderClass.API_JSON_DATA_FILE, 'r')
                json_out = json.load(jfile)
                for k, v in json_out.items():
                    DataProviderClass.TEST_DATA_DICT[k] = v
        except FileNotFoundError:
            print("No Such File Found {}".format(DataProviderClass.JSON_DATA_FILE))
        finally:
            jfile.close()

    @staticmethod
    def get_data_for_testcase(testcase, source, type="sel"):
        if testcase in DataProviderClass.TEST_DATA_DICT:
            new_data = DataProviderClass.TEST_DATA_DICT
            return DataProviderClass.TEST_DATA_DICT[testcase]
        else:
            if source == "excel":
                DataProviderClass.populate_data_from_excel(type)
            elif source == "json":
                DataProviderClass.populate_data_from_json(type)
            else:
                DataProviderClass.populate_data_from_local(type)
        return DataProviderClass.TEST_DATA_DICT[testcase]

